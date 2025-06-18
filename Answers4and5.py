from openpyxl import load_workbook
import math

wb = load_workbook("sagatave_eksamenam .xlsx", data_only=True)
ws = wb["Lapa_0"]

headers = [cell.value for cell in ws[3]]
columns = {header: idx for idx, header in enumerate(headers)}

total_cena = 0
count_cena = 0
total_kopa = 0

for row in ws.iter_rows(min_row=4, values_only=True):
    produkts = row[columns["Produkts"]]
    cena = row[columns["Cena"]]
    klients = row[columns["Klients"]]
    skaits = row[columns["Skaits"]]
    kopa = row[columns["Kopā"]]

    if "LaserJet" in str(produkts):
        total_cena += cena
        count_cena += 1

    if klients == "Korporatīvais" and skaits is not None and 40 <= skaits <= 50:
        total_kopa += kopa

average_cena = math.floor(total_cena / count_cena) if count_cena else 0
total_kopa = math.floor(total_kopa)

print("Question 4 Answer (Average Cena for LaserJet):", average_cena)
print("Question 5 Answer (Total Kopā for Korporatīvais, Skaits 40–50):", total_kopa)
